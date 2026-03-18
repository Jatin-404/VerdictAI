from fastapi import APIRouter, BackgroundTasks, UploadFile, File, Form, HTTPException
from pydantic import BaseModel
from httpx import AsyncClient
import uuid
import asyncio
import json
import zipfile
import io
import os
import traceback

# PDF and DOCX readers
import fitz                          # pymupdf -> for PDFs
from docx import Document as DocxDoc # python-docx -> for DOCX


from app.core.config import settings

CHUNKER_URL = settings.CHUNKER_URL
EMBED_URL   = settings.EMBED_URL
STORE_URL   = settings.STORE_URL


router = APIRouter(prefix='/ingest', tags=["Ingestion"])


jobs = {}

# MODELS

class CaseMetadata(BaseModel):
    judgment_id:   str      # "IN-HC-ALL-2006-CV-121D60"
    court:         str      # "Allahabad High Court"
    court_level:   str      # "HC" or "SC"
    decision_date: str      # "27 MARCH, 2006"
    domain:        str      # "service" / "civil" / "criminal"
    bench:         str      # judge name
    jurisdiction:  str      # "India"

class IngestRequest(BaseModel):
    content_text: str
    metadata: CaseMetadata
    document_id: str | None = None
    filename: str
    chunk_size: int = 800
    overlap: int = 200      # ← add this

# FILE TEXT EXTRACTION

def extract_text_from_file(filename: str, file_bytes: bytes) -> str:
    """
    Reads txt, pdf, or docx and returns plain text.
    This is why we installed pymupdf and python-docx.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".txt":
        return file_bytes.decode("utf-8")

    elif ext == ".pdf":
        # fitz opens from bytes in memory - no temp file needed
        pdf = fitz.open(stream=file_bytes, filetype = "pdf")
        text = ''
        for page in pdf:
            text += page.get_text()

        return text.strip()
    
    elif ext == ".docx":
        # python-docx needs a file-like obejct
        doc = DocxDoc(io.BytesIO(file_bytes))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text.strip()
    
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .txt, .pdf, .docx")
    
def parse_metadata(metadata_bytes: bytes, filename: str) -> CaseMetadata:
    """
    Reads metadata from either JSON or XLSX.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".json":
        try:
            data = json.loads(metadata_bytes.decode("utf-8"))
            # lawyers might come as comma-separated string from Excel
            if isinstance(data.get("lawyers"), str):
                data["lawyers"] = [l.strip() for l in data["lawyers"].split(",")]
            return CaseMetadata(**data)
        except Exception as e:
            raise ValueError(f"Invalid metadata JSON: {str(e)}")
        
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(metadata_bytes))
        ws = wb.active
        # First row = headers, secod row = values
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]
        data = dict(zip(headers,values))

        # lawyers column: "Advocate A, Advocate B" -> list

        if isinstance(data.get("lawyers"), str):
            data["lawyers"] = [l.strip() for l in data["lawyers"].split(",")]

        return CaseMetadata(**data)
    
    else:
        raise ValueError(f"Unsupported metadata format: {ext}. Use .json or .xlsx")




# Core Pipeline


async def ingest_one(request: IngestRequest, client: AsyncClient):

    chunk_response = await client.post(f"{CHUNKER_URL}/chunk",
                                        json={
                                            "content_text": request.content_text,
                                            "metadata": request.metadata.model_dump(),
                                            "filename": request.filename,
                                            "chunk_size": request.chunk_size,
                                            "overlap": request.overlap
                                        }) 
    chunk_data = chunk_response.json()
    chunks = chunk_data["chunks"]
    document_id = chunk_data["document_id"]

    embed_batch_size = 20
    all_embedded_chunks = []

    for i in range(0, len(chunks), embed_batch_size):
        batch = chunks[i: i + embed_batch_size]
        embed_response = await client.post(
            f"{EMBED_URL}/embed",
            json={"chunks": batch}
        )
        embedded = embed_response.json()["chunks"]
        all_embedded_chunks.extend(embedded)



    await client.post(f"{STORE_URL}/store",
                        json={"chunks": all_embedded_chunks})

    return {
        "document_id": document_id,
        "total_chunks": len(chunks),
        "judgment_id": request.metadata.judgment_id,
        "court": request.metadata.court
    }
    
async def run_ingest_batch(requests: list[IngestRequest], job_id: str):
    jobs[job_id] = {"status": "processing"}
    completed = []
    failed = []
    try:
        batch_size = 10
        async with AsyncClient(timeout=600) as client:
            for i in range(0, len(requests), batch_size):
                batch = requests[i : i + batch_size]
                tasks = [ingest_one(req, client) for req in batch]
                results = await asyncio.gather(*tasks, return_exceptions=True)

                # ← MUST be inside the for loop (indented one more level)
                for req, result in zip(batch, results):
                    if isinstance(result, Exception):
                        failed.append({
                            "judgment_id": req.metadata.judgment_id,
                            "error": str(result) if str(result) else repr(result),
                            "type": type(result).__name__
                        })
                    else:
                        completed.append(result)

                # ← progress update also inside the for loop
                jobs[job_id] = {
                    "status": "processing",
                    "total_cases": len(requests),
                    "succeeded_so_far": len(completed),
                    "failed_so_far": len(failed),
                    "progress": f"{i + len(batch)}/{len(requests)}"
                }

        # ← final status OUTSIDE the for loop and with block
        jobs[job_id] = {
            "status": "completed",
            "total_cases": len(requests),
            "succeeded": len(completed),
            "failed": len(failed),
            "results": completed,
            "errors": failed
        }

    except Exception as e:
        jobs[job_id] = {"status": "failed", "error": repr(e)}


def parse_json_zip(zip_bytes: bytes, chunk_size: int, overlap: int):
    valid_requests = []
    errors = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for path in zf.namelist():
            if not path.endswith(".json"):
                continue
            try:
                raw = zf.read(path)
                data = json.loads(raw.decode("utf-8"))

                # extract text
                content_text = data.get("text", "").strip()
                if not content_text:
                    raise ValueError("No text field found")

                # extract metadata
                meta = data.get("metadata", {})
                classification = data.get("classification", {})

                metadata = CaseMetadata(
                    judgment_id   = data.get("judgment_id", path),
                    court         = meta.get("court", "Unknown"),
                    court_level   = meta.get("court_level", "Unknown"),
                    decision_date = meta.get("decision_date", "Unknown"),
                    domain        = classification.get("domain", "Unknown"),
                    bench         = meta.get("bench", "")[:200],  # bench field is very long, truncate
                    jurisdiction  = meta.get("jurisdiction", "India")
                )

                valid_requests.append(IngestRequest(
                    content_text=content_text,
                    metadata=metadata,
                    filename=os.path.basename(path),
                    chunk_size=chunk_size,
                    overlap=overlap
                ))

            except Exception as e:
                errors.append({"file": path, "error": str(e)})

    return valid_requests, errors




# ENDPOINTS

@router.post('/')
async def ingest(requests: list[IngestRequest], background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}

    background_tasks.add_task(run_ingest_batch, requests, job_id)
    return{
        "job_id": job_id,
        "status": "queued",
        "msg": "ingestion started in background"
    }

@router.post("/upload")
async def ingest_upload(
    background_tasks: BackgroundTasks,
    content_file: UploadFile = File(...),
    metadata_file: UploadFile = File(...),
    chunk_size: int = Form(default=500)  # optional sent as form field
    ):
    # step-1 reading both files 
    content_bytes = await content_file.read()           # without await  we would get a coroutine object here 
    metadata_bytes = await metadata_file.read()         # the info stored after reading is in bytes

    # step-2 decode content file as plain text

    try: 
        content_text = extract_text_from_file(content_file.filename, content_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail= str(e))
    
    # step-3 parsing metadata file as json

    try:
        metadata = parse_metadata(metadata_bytes, metadata_file.filename)
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail= str(e))
       
    
    # step-4 Build ingest request (same shape as before)

    request = IngestRequest(
        content_text=content_text,
        metadata=metadata,
        filename=content_file.filename,
        chunk_size=chunk_size
    )

    # step 5 handoff to same background pipeline

    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}
    background_tasks.add_task(run_ingest_batch, [request], job_id)

    return{
        "job_id": job_id,
        "status": "queued",
        "filename": content_file.filename,
        "metadata_file": metadata_file.filename,
        "msg": "File ingestion started in background"  
    }

# ZIP BATCH UPLOAD (multiple cases at once)
@router.post("/batch")
async def ingest_batch_zip(
        background_tasks: BackgroundTasks,
        zip_file: UploadFile = File(...),
        chunk_size: int = Form(default=800),
        overlap: int = Form(default=200)
    ):
    # validate its actually a zip
    if not zip_file.filename.endswith(".zip"):
        raise HTTPException(status_code=400, detail="File must be a .zip")

    zip_bytes = await zip_file.read()

    # Parse zip -> list of IngestRequests
    try:
        valid_requests, parse_errors = parse_json_zip(zip_bytes, chunk_size, overlap)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read zip: {str(e)}")
    
    if not valid_requests:
        raise HTTPException(
            status_code=400,
            detail=f"No valid cases found in zip. Errors: {parse_errors}"
        )
    
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}
    background_tasks.add_task(run_ingest_batch, valid_requests,job_id)

    return{
        "job_id": job_id,
        "status": "queued",
        "total_cases_found": len(valid_requests),
        "skipped_folders": parse_errors,   # folders that had missing/bad files
        "msg": "ZIP batch ingestion started in background"
    }


@router.get('/jobs/{job_id}')
async def get_job(job_id : str):
    if job_id not in jobs:
        return{"error": "job not found"}
    else:
        return{
            "job_id": job_id,
            **jobs[job_id]
        }
    